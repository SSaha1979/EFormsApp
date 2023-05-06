import pandas as pd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
# Load the data from the Excel sheet
filename = "C:\\WEIBULL\\FOR CHARTS.xlsx"
# # # Load/ open excel file
workbook = openpyxl.load_workbook(filename)
# sheet = workbook["output"]


# filename = "C:\\SHWETA\\UserID.xlsx"
# # Load/ open excel file
# workbook = openpyxl.load_workbook(filename)
# sheet = workbook["UserID"]



# df = pd.read_excel('C:/Users/Admin/Desktop/Shweta111/FOR CHART.xlsx')
# Calculate the probability of failure (p) for each value of SF
p = 1 - sheet.cell(column=4).value
# p = 1 - df['SF']
# Define the Weibull function
def weibull(x, beta, eta):
    return (beta/eta) * ((x/eta)**(beta-1)) * np.exp(-(x/eta)**beta)
# Fit the Weibull function to the data using curve_fit
params,cov=curve_fit(weibull, sheet.cell(column=0), p)
params,cov=curve_fit
# Extract the Weibull parameters beta and eta
beta = params[0]
eta = params[1]
# Calculate the fitted probabilities using the Weibull function
p_fit = weibull(sheet.cell['MIS'], beta, eta)
# Plot the data and the Weibull fit
fig, ax = plt.subplots()
ax.scatter(sheet.cell(column=0), p)
ax.plot(sheet.cell(column=0), p_fit, color='red')
# Set the axis labels and title
ax.set_xlabel('MIS')
ax.set_ylabel('SF')
ax.set_title('Weibull Graph')
# Display the plot
plt.show()












